"""
services/excel_service.py
==========================
LOGIPORT — Excel Export Service  (openpyxl-based)

Supported exports:
    export_transactions(...)       → Transactions list with totals
    export_transaction_items(...)  → Full item breakdown for one transaction
    export_materials(...)          → Materials catalogue
    export_clients(...)            → Clients list
    export_pricing(...)            → Pricing table
"""
from __future__ import annotations

import logging
import re
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False
    logger.warning("openpyxl not installed — Excel export unavailable")


# ── Styling ───────────────────────────────────────────────────────────────────
_HEADER_BG  = "1A3A5C"
_TOTAL_BG   = "EFF6FF"
_ALT_BG     = "F8FAFC"
_BORDER_CLR = "CBD5E1"

_thin   = lambda: Side(style="thin", color=_BORDER_CLR)
_border = lambda: Border(left=_thin(), right=_thin(), top=_thin(), bottom=_thin())

def _hf(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, name="Calibri", size=size)

_HEADER_FILL = lambda: PatternFill("solid", fgColor=_HEADER_BG)
_TOTAL_FILL  = lambda: PatternFill("solid", fgColor=_TOTAL_BG)
_ALT_FILL    = lambda: PatternFill("solid", fgColor=_ALT_BG)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
_RIGHT  = Alignment(horizontal="right",  vertical="center")


def _style_header(ws, row: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font      = _hf(bold=True, color="FFFFFF", size=11)
        cell.fill      = _HEADER_FILL()
        cell.border    = _border()
        cell.alignment = _CENTER

def _style_data(ws, r0: int, r1: int, ncols: int):
    for r in range(r0, r1 + 1):
        fill = _ALT_FILL() if r % 2 == 0 else None
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font      = _hf()
            cell.border    = _border()
            cell.alignment = _LEFT
            if fill:
                cell.fill = fill

def _style_total(ws, row: int, ncols: int):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font      = _hf(bold=True, color="1E40AF")
        cell.fill      = _TOTAL_FILL()
        cell.border    = _border()
        cell.alignment = _RIGHT

def _write_title(ws, title: str, subtitle: str = "") -> int:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
    c = ws.cell(row=1, column=1, value=title)
    c.font = _hf(bold=True, color="1A3A5C", size=14)
    c.alignment = _CENTER
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
        c2 = ws.cell(row=2, column=1, value=subtitle)
        c2.font = _hf(color="64748B", size=10)
        c2.alignment = _CENTER
        return 4
    return 3


# ── Helpers ───────────────────────────────────────────────────────────────────
def _require_openpyxl():
    if not _HAS_OPENPYXL:
        raise ImportError("openpyxl is required. Install: pip install openpyxl")

def _export_dir() -> Path:
    try:
        from core.settings_manager import SettingsManager
        custom = SettingsManager.get_instance().get_documents_output_path()
        base = Path(custom) if custom else Path("documents") / "output"
    except Exception:
        base = Path("documents") / "output"
    out = base / "exports"
    out.mkdir(parents=True, exist_ok=True)
    return out

def _ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def _s(v: Any) -> str:
    return "" if v is None else str(v)

def _rel(obj: Any, lang: str = "en") -> str:
    if obj is None:
        return ""
    for attr in (f"name_{lang}", "name_en", "name_ar", "name", "code"):
        val = getattr(obj, attr, None)
        if val:
            return str(val)
    return str(obj)

def _session():
    from database.models import get_session_local
    return get_session_local()()


# ── Column definitions: (key, header_ar, header_en, header_tr, width) ─────────
_TRX_COLS = [
    ("transaction_no",   "رقم المعاملة",      "Trx No",        "İşlem No",           16),
    ("transaction_date", "التاريخ",            "Date",          "Tarih",              14),
    ("transaction_type", "النوع",              "Type",          "Tür",                12),
    ("client",           "الزبون",             "Client",        "Müşteri",            22),
    ("exporter_company", "المصدِّر",           "Exporter",      "İhracatçı",          22),
    ("importer_company", "المستورِد",          "Importer",      "İthalatçı",          22),
    ("origin_country",   "بلد المنشأ",         "Origin",        "Menşei",             14),
    ("dest_country",     "بلد المقصد",         "Destination",   "Varış",              14),
    ("currency",         "العملة",             "Currency",      "Para Birimi",        10),
    ("delivery_method",  "طريقة التسليم",      "Delivery",      "Teslimat",           16),
    ("transport_type",   "وسيلة النقل",        "Transport",     "Taşıma",             12),
    ("transport_ref",    "مرجع النقل",         "Transport Ref", "Taşıma Ref",         16),
    ("totals_count",     "عدد الأصناف",        "Items",         "Öğe Sayısı",         10),
    ("totals_gross_kg",  "الوزن القائم (كغ)",  "Gross (kg)",    "Brüt (kg)",          14),
    ("totals_net_kg",    "الوزن الصافي (كغ)",  "Net (kg)",      "Net (kg)",           14),
    ("totals_value",     "القيمة الإجمالية",   "Total Value",   "Toplam Değer",       16),
    ("notes",            "ملاحظات",            "Notes",         "Notlar",             30),
]

_ITEM_COLS = [
    ("seq",              "م",                  "#",             "#",                   6),
    ("material",         "المادة",             "Material",      "Malzeme",            24),
    ("packaging_type",   "نوع التغليف",        "Packaging",     "Ambalaj",            16),
    ("quantity",         "الكمية",             "Quantity",      "Miktar",             12),
    ("gross_weight_kg",  "الوزن القائم (كغ)",  "Gross (kg)",    "Brüt (kg)",          14),
    ("net_weight_kg",    "الوزن الصافي (كغ)",  "Net (kg)",      "Net (kg)",           14),
    ("unit_price",       "سعر الوحدة",         "Unit Price",    "Birim Fiyat",        14),
    ("currency",         "العملة",             "Currency",      "Para Birimi",        10),
    ("line_total",       "الإجمالي",           "Line Total",    "Toplam",             14),
    ("origin_country",   "بلد المنشأ",         "Origin",        "Menşei",             14),
    ("transport_ref",    "مرجع النقل",         "Transport Ref", "Taşıma Ref",         16),
    ("notes",            "ملاحظات",            "Notes",         "Notlar",             30),
]

_MAT_COLS = [
    ("id",            "الرقم",               "ID",         "ID",            8),
    ("name_ar",       "الاسم بالعربية",       "Name (AR)",  "Ad (AR)",       24),
    ("name_en",       "الاسم بالإنجليزية",   "Name (EN)",  "Ad (EN)",       24),
    ("name_tr",       "الاسم بالتركية",      "Name (TR)",  "Ad (TR)",       24),
    ("code",          "الرمز",               "Code",       "Kod",           14),
    ("material_type", "نوع المادة",          "Type",       "Tür",           18),
    ("notes",         "ملاحظات",             "Notes",      "Notlar",        30),
]

_CLIENT_COLS = [
    ("id",      "الرقم",               "ID",      "ID",         8),
    ("name",    "الاسم",               "Name",    "Ad",         24),
    ("country", "الدولة",             "Country", "Ülke",       14),
    ("phone",   "الهاتف",             "Phone",   "Telefon",    16),
    ("email",   "البريد الإلكتروني",  "Email",   "E-posta",    26),
    ("address", "العنوان",            "Address", "Adres",      30),
    ("notes",   "ملاحظات",            "Notes",   "Notlar",     30),
]

_PRICING_COLS = [
    ("id",              "الرقم",           "ID",           "ID",            8),
    ("seller_company",  "البائع",          "Seller",       "Satıcı",        22),
    ("buyer_company",   "المشتري",         "Buyer",        "Alıcı",         22),
    ("material",        "المادة",          "Material",     "Malzeme",       24),
    ("pricing_type",    "نوع التسعيرة",    "Price Type",   "Fiyat Tipi",    16),
    ("unit_price",      "سعر الوحدة",      "Unit Price",   "Birim Fiyat",   14),
    ("currency",        "العملة",          "Currency",     "Para Birimi",   10),
    ("delivery_method", "طريقة التسليم",   "Delivery",     "Teslimat",      16),
    ("is_active",       "نشطة",            "Active",       "Aktif",         10),
    ("notes",           "ملاحظات",         "Notes",        "Notlar",        30),
]

_LANG_IDX = {"ar": 1, "en": 2, "tr": 3}

def _hdr(col_def: tuple, lang: str) -> str:
    return col_def[_LANG_IDX.get(lang, 2)]

def _wid(col_def: tuple) -> int:
    return col_def[4]


# ═══════════════════════════════════════════════════════════════════════════════
class ExcelService:
    """Builds .xlsx exports from LOGIPORT database data."""

    # ── transactions ──────────────────────────────────────────────────────────
    def export_transactions(
        self, *,
        date_from: Optional[str] = None,
        date_to:   Optional[str] = None,
        client_id: Optional[int] = None,
        transaction_type: Optional[str] = None,
        lang: str = "ar",
        output_path: Optional[Path] = None,
    ) -> Path:
        """Export transactions list. Returns Path to generated .xlsx."""
        _require_openpyxl()
        rows = self._fetch_transactions(date_from, date_to, client_id, transaction_type, lang)

        wb = Workbook()
        ws = wb.active
        ws.title = "المعاملات" if lang == "ar" else "Transactions"
        ws.sheet_view.rightToLeft = (lang == "ar")

        parts = []
        if date_from: parts.append(f"من: {date_from}")
        if date_to:   parts.append(f"إلى: {date_to}")
        ds = _write_title(ws, "LOGIPORT — تصدير المعاملات", "  |  ".join(parts))

        self._write_sheet(ws, _TRX_COLS, rows, ds, lang)
        self._add_totals(ws, _TRX_COLS, rows, ds,
                         {"totals_count","totals_gross_kg","totals_net_kg","totals_value"}, lang)

        out = output_path or (_export_dir() / f"transactions_{_ts()}.xlsx")
        wb.save(out)
        logger.info("Transactions exported → %s (%d rows)", out, len(rows))
        return Path(out)

    # ── transaction items ─────────────────────────────────────────────────────
    def export_transaction_items(
        self,
        transaction_id: int, *,
        lang: str = "ar",
        output_path: Optional[Path] = None,
    ) -> Path:
        """Export full item breakdown for a single transaction."""
        _require_openpyxl()
        header_info, items = self._fetch_items(transaction_id, lang)

        wb = Workbook()
        ws = wb.active
        ws.title = "البنود" if lang == "ar" else "Items"
        ws.sheet_view.rightToLeft = (lang == "ar")

        tx_no   = header_info.get("transaction_no", str(transaction_id))
        tx_date = header_info.get("transaction_date", "")
        ds = _write_title(ws, f"LOGIPORT — معاملة {tx_no}", f"التاريخ: {tx_date}")

        # mini header block
        info = [
            ("الزبون" if lang=="ar" else "Client",    header_info.get("client","")),
            ("المصدِّر" if lang=="ar" else "Exporter", header_info.get("exporter_company","")),
            ("المستورِد" if lang=="ar" else "Importer",header_info.get("importer_company","")),
            ("العملة"  if lang=="ar" else "Currency",  header_info.get("currency","")),
        ]
        for i, (lbl, val) in enumerate(info):
            ws.cell(row=ds+i, column=1, value=lbl).font = _hf(bold=True)
            ws.cell(row=ds+i, column=2, value=val).font = _hf()
        ds += len(info) + 1

        self._write_sheet(ws, _ITEM_COLS, items, ds, lang)
        self._add_totals(ws, _ITEM_COLS, items, ds,
                         {"quantity","gross_weight_kg","net_weight_kg","line_total"}, lang)

        out = output_path or (_export_dir() / f"trx_{tx_no}_{_ts()}.xlsx")
        wb.save(out)
        logger.info("Items exported → %s (%d items)", out, len(items))
        return Path(out)

    # ── materials ─────────────────────────────────────────────────────────────
    def export_materials(self, *, lang: str = "ar", output_path: Optional[Path] = None) -> Path:
        _require_openpyxl()
        rows = self._fetch_materials(lang)
        return self._simple_export(rows, _MAT_COLS, lang,
                                   "LOGIPORT — المواد", "materials", output_path)

    # ── clients ───────────────────────────────────────────────────────────────
    def export_clients(self, *, lang: str = "ar", output_path: Optional[Path] = None) -> Path:
        _require_openpyxl()
        rows = self._fetch_clients(lang)
        return self._simple_export(rows, _CLIENT_COLS, lang,
                                   "LOGIPORT — الزبائن", "clients", output_path)

    # ── pricing ───────────────────────────────────────────────────────────────
    def export_pricing(self, *, lang: str = "ar", output_path: Optional[Path] = None) -> Path:
        _require_openpyxl()
        rows = self._fetch_pricing(lang)
        return self._simple_export(rows, _PRICING_COLS, lang,
                                   "LOGIPORT — التسعيرة", "pricing", output_path)

    # ── shared sheet writer ───────────────────────────────────────────────────
    def _simple_export(self, rows, cols, lang, title, fname_prefix, output_path) -> Path:
        wb = Workbook()
        ws = wb.active
        ws.sheet_view.rightToLeft = (lang == "ar")
        ds = _write_title(ws, title)
        self._write_sheet(ws, cols, rows, ds, lang)
        out = output_path or (_export_dir() / f"{fname_prefix}_{_ts()}.xlsx")
        wb.save(out)
        logger.info("%s exported → %s (%d rows)", fname_prefix, out, len(rows))
        return Path(out)

    def _write_sheet(self, ws, cols, rows, data_start: int, lang: str):
        ncols = len(cols)
        # headers
        for ci, col in enumerate(cols, 1):
            ws.cell(row=data_start, column=ci, value=_hdr(col, lang))
        _style_header(ws, data_start, ncols)

        # data
        for ri, row in enumerate(rows, data_start + 1):
            for ci, col in enumerate(cols, 1):
                ws.cell(row=ri, column=ci, value=row.get(col[0], ""))
        if rows:
            _style_data(ws, data_start + 1, data_start + len(rows), ncols)

        # column widths & freeze
        for ci, col in enumerate(cols, 1):
            ws.column_dimensions[get_column_letter(ci)].width = _wid(col)
        ws.freeze_panes = ws.cell(row=data_start + 1, column=1)

    def _add_totals(self, ws, cols, rows, data_start: int, sum_keys: set, lang: str):
        if not rows:
            return
        ncols  = len(cols)
        tr     = data_start + len(rows) + 1
        labels = {"ar": "الإجمالي", "en": "Total", "tr": "Toplam"}
        ws.cell(row=tr, column=1, value=labels.get(lang, "Total"))
        for ci, col in enumerate(cols, 1):
            if col[0] in sum_keys:
                total = sum(float(r.get(col[0]) or 0) for r in rows)
                ws.cell(row=tr, column=ci, value=round(total, 4))
        _style_total(ws, tr, ncols)

    # ── data fetchers ─────────────────────────────────────────────────────────
    def _fetch_transactions(self, date_from, date_to, client_id, trx_type, lang) -> List[Dict]:
        from sqlalchemy import select
        from sqlalchemy.orm import joinedload
        from database.models import Transaction, TransactionItem

        with _session() as s:
            q = (select(Transaction)
                 .options(
                     joinedload(Transaction.client),
                     joinedload(Transaction.exporter_company),
                     joinedload(Transaction.importer_company),
                     joinedload(Transaction.origin_country),
                     joinedload(Transaction.dest_country),
                     joinedload(Transaction.currency),
                     joinedload(Transaction.delivery_method),
                 )
                 .order_by(Transaction.transaction_date.desc(), Transaction.id.desc()))
            if client_id:   q = q.where(Transaction.client_id == client_id)
            if trx_type:    q = q.where(Transaction.transaction_type == trx_type)
            if date_from:   q = q.where(Transaction.transaction_date >= date_from)
            if date_to:     q = q.where(Transaction.transaction_date <= date_to)
            txs = list(s.execute(q).unique().scalars())
            s.expunge_all()

        return [{
            "transaction_no":   _s(t.transaction_no),
            "transaction_date": _s(t.transaction_date),
            "transaction_type": _s(t.transaction_type),
            "client":           _rel(t.client, lang),
            "exporter_company": _rel(t.exporter_company, lang),
            "importer_company": _rel(t.importer_company, lang),
            "origin_country":   _rel(t.origin_country, lang),
            "dest_country":     _rel(t.dest_country, lang),
            "currency":         _rel(t.currency, lang),
            "delivery_method":  _rel(t.delivery_method, lang),
            "transport_type":   _s(t.transport_type),
            "transport_ref":    _s(t.transport_ref),
            "totals_count":     float(t.totals_count or 0),
            "totals_gross_kg":  float(t.totals_gross_kg or 0),
            "totals_net_kg":    float(t.totals_net_kg or 0),
            "totals_value":     float(t.totals_value or 0),
            "notes":            _s(t.notes),
        } for t in txs]

    def _fetch_items(self, trx_id: int, lang: str) -> Tuple[Dict, List[Dict]]:
        from sqlalchemy import select
        from sqlalchemy.orm import joinedload
        from database.models import Transaction, TransactionItem

        with _session() as s:
            trx = s.execute(
                select(Transaction)
                .where(Transaction.id == trx_id)
                .options(
                    joinedload(Transaction.client),
                    joinedload(Transaction.exporter_company),
                    joinedload(Transaction.importer_company),
                    joinedload(Transaction.currency),
                    joinedload(Transaction.items).joinedload(TransactionItem.material),
                    joinedload(Transaction.items).joinedload(TransactionItem.packaging_type),
                    joinedload(Transaction.items).joinedload(TransactionItem.currency),
                    joinedload(Transaction.items).joinedload(TransactionItem.origin_country),
                )
            ).unique().scalar_one_or_none()
            if not trx:
                return {}, []

            header = {
                "transaction_no":   _s(trx.transaction_no),
                "transaction_date": _s(trx.transaction_date),
                "transaction_type": _s(trx.transaction_type),
                "client":           _rel(trx.client, lang),
                "exporter_company": _rel(trx.exporter_company, lang),
                "importer_company": _rel(trx.importer_company, lang),
                "currency":         _rel(trx.currency, lang),
            }
            items_data = [{
                "seq":             seq,
                "material":        _rel(it.material, lang),
                "packaging_type":  _rel(it.packaging_type, lang),
                "quantity":        float(it.quantity or 0),
                "gross_weight_kg": float(it.gross_weight_kg or 0),
                "net_weight_kg":   float(it.net_weight_kg or 0),
                "unit_price":      float(it.unit_price or 0),
                "currency":        _rel(it.currency, lang),
                "line_total":      float(it.line_total or 0),
                "origin_country":  _rel(it.origin_country, lang),
                "transport_ref":   _s(it.transport_ref),
                "notes":           _s(it.notes),
            } for seq, it in enumerate(trx.items, 1)]
            s.expunge_all()
        return header, items_data

    def _fetch_materials(self, lang: str) -> List[Dict]:
        from sqlalchemy import select
        from sqlalchemy.orm import joinedload
        from database.models import Material

        with _session() as s:
            mats = list(s.execute(
                select(Material).order_by(Material.id)
            ).unique().scalars())
            rows = []
            for m in mats:
                mt = ""
                if hasattr(m, "material_type") and m.material_type:
                    mt = _rel(m.material_type, lang)
                rows.append({
                    "id":            m.id,
                    "name_ar":       _s(getattr(m, "name_ar", "")),
                    "name_en":       _s(getattr(m, "name_en", "")),
                    "name_tr":       _s(getattr(m, "name_tr", "")),
                    "code":          _s(getattr(m, "code", "")),
                    "material_type": mt,
                    "notes":         _s(getattr(m, "notes", "")),
                })
            s.expunge_all()
        return rows

    def _fetch_clients(self, lang: str) -> List[Dict]:
        from sqlalchemy import select
        from sqlalchemy.orm import joinedload
        from database.models import Client

        with _session() as s:
            clients = list(s.execute(
                select(Client).order_by(Client.id)
            ).unique().scalars())
            rows = []
            for c in clients:
                country = ""
                if hasattr(c, "country") and c.country:
                    country = _rel(c.country, lang)
                rows.append({
                    "id":      c.id,
                    "name":    _s(c.name),
                    "country": country,
                    "phone":   _s(getattr(c, "phone", "")),
                    "email":   _s(getattr(c, "email", "")),
                    "address": _s(getattr(c, "address", "")),
                    "notes":   _s(getattr(c, "notes", "")),
                })
            s.expunge_all()
        return rows

    def _fetch_pricing(self, lang: str) -> List[Dict]:
        from sqlalchemy import select
        from sqlalchemy.orm import joinedload
        try:
            from database.models import Pricing
        except ImportError:
            return []

        with _session() as s:
            try:
                pricings = list(s.execute(select(Pricing).order_by(Pricing.id)).unique().scalars())
            except Exception:
                return []
            rows = [{
                "id":              p.id,
                "seller_company":  _rel(getattr(p, "seller_company", None), lang),
                "buyer_company":   _rel(getattr(p, "buyer_company", None), lang),
                "material":        _rel(getattr(p, "material", None), lang),
                "pricing_type":    _rel(getattr(p, "pricing_type", None), lang),
                "unit_price":      float(p.unit_price or 0),
                "currency":        _rel(getattr(p, "currency", None), lang),
                "delivery_method": _rel(getattr(p, "delivery_method", None), lang),
                "is_active":       "✓" if getattr(p, "is_active", True) else "✗",
                "notes":           _s(getattr(p, "notes", "")),
            } for p in pricings]
            s.expunge_all()
        return rows