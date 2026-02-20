"""
tests/test_documents.py
========================
Covers:
  TestExcelService          — Excel export pure logic (no DB)
  TestStorageService        — Path building and sanitization
  TestNumberingServicePure  — Pure helper functions (no DB)
  TestDocumentsDB           — doc_groups and documents CRUD via in-memory DB
"""
import pytest
import tempfile
import re
from pathlib import Path
from datetime import date
from unittest.mock import patch, MagicMock


# ══════════════════════════════════════════════════════════════════════════════
# Excel Service — pure sheet writing logic (no DB needed)
# ══════════════════════════════════════════════════════════════════════════════

class TestExcelService:
    """Tests for services/excel_service.py — writing logic only."""

    @pytest.fixture(autouse=True)
    def _load(self):
        # Load in isolation (no DB chain)
        import importlib.util
        spec = importlib.util.spec_from_file_location(
            "excel_service",
            Path(__file__).parent.parent / "services" / "excel_service.py",
        )
        self.mod = importlib.util.load_module = None
        ns = {}
        exec(
            compile(
                (Path(__file__).parent.parent / "services" / "excel_service.py").read_text(),
                "excel_service.py",
                "exec",
            ),
            ns,
        )
        self.ExcelService = ns["ExcelService"]
        self._write_title = ns["_write_title"]
        self._style_header = ns["_style_header"]
        self._style_data = ns["_style_data"]
        self._style_total = ns["_style_total"]
        self._TRX_COLS = ns["_TRX_COLS"]
        self._MAT_COLS = ns["_MAT_COLS"]
        self._CLIENT_COLS = ns["_CLIENT_COLS"]

    # ── _safe_str ─────────────────────────────────────────────────────────────

    def test_s_helper_none(self):
        ns = {}
        exec(
            compile(
                (Path(__file__).parent.parent / "services" / "excel_service.py").read_text(),
                "excel_service.py", "exec"
            ),
            ns,
        )
        assert ns["_s"](None) == ""
        assert ns["_s"]("hello") == "hello"
        assert ns["_s"](123) == "123"

    # ── Workbook writing ──────────────────────────────────────────────────────

    def test_write_title_returns_next_row(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        row = self._write_title(ws, "Test Title")
        assert row >= 3  # at least row 3 (title + blank)

    def test_write_title_with_subtitle(self):
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        row = self._write_title(ws, "Main Title", "Sub Title")
        assert row >= 4

    def test_write_sheet_creates_header_row(self):
        import openpyxl
        svc = self.ExcelService()
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = [{"transaction_no": "T001", "transaction_date": "2026-01-01",
                 "transaction_type": "export", "client": "Test"}]
        svc._write_sheet(ws, self._TRX_COLS[:4], rows, 1, "en")
        assert ws.cell(row=1, column=1).value == "Trx No"
        assert ws.cell(row=2, column=1).value == "T001"

    def test_write_sheet_arabic_headers(self):
        import openpyxl
        svc = self.ExcelService()
        wb = openpyxl.Workbook()
        ws = wb.active
        svc._write_sheet(ws, self._TRX_COLS[:4], [], 1, "ar")
        assert ws.cell(row=1, column=1).value == "رقم المعاملة"

    def test_write_sheet_turkish_headers(self):
        import openpyxl
        svc = self.ExcelService()
        wb = openpyxl.Workbook()
        ws = wb.active
        svc._write_sheet(ws, self._TRX_COLS[:4], [], 1, "tr")
        assert ws.cell(row=1, column=1).value == "İşlem No"

    def test_add_totals_sums_numeric_cols(self):
        import openpyxl
        svc = self.ExcelService()
        wb = openpyxl.Workbook()
        ws = wb.active
        cols = self._TRX_COLS
        rows = [
            {"totals_value": 1000.0, "totals_gross_kg": 500.0, "totals_net_kg": 450.0,
             "totals_count": 3.0},
            {"totals_value": 2000.0, "totals_gross_kg": 800.0, "totals_net_kg": 750.0,
             "totals_count": 5.0},
        ]
        svc._write_sheet(ws, cols, rows, 1, "en")
        svc._add_totals(ws, cols, rows, 1,
                        {"totals_count", "totals_gross_kg", "totals_net_kg", "totals_value"},
                        "en")
        total_row = 1 + len(rows) + 1
        # Find totals_value col index
        val_idx = next(i+1 for i, c in enumerate(cols) if c[0] == "totals_value")
        total_val = ws.cell(row=total_row, column=val_idx).value
        assert total_val == 3000.0

    def test_export_to_temp_file(self):
        """Full export cycle writes a valid .xlsx file."""
        import openpyxl
        svc = self.ExcelService()
        wb = openpyxl.Workbook()
        ws = wb.active
        svc._write_sheet(ws, self._MAT_COLS, [
            {"id": 1, "name_ar": "قمح", "name_en": "Wheat", "name_tr": "Buğday",
             "code": "WHT", "material_type": "", "notes": ""}
        ], 1, "ar")
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "test_materials.xlsx"
            wb.save(path)
            assert path.exists()
            wb2 = openpyxl.load_workbook(path)
            assert wb2.active.max_row >= 2


# ══════════════════════════════════════════════════════════════════════════════
# Storage Service — path building
# ══════════════════════════════════════════════════════════════════════════════

class TestStorageService:

    @pytest.fixture(autouse=True)
    def _import(self):
        import importlib.util
        ns = {}
        exec(
            compile(
                (Path(__file__).parent.parent / "services" / "storage_service.py").read_text(),
                "storage_service.py", "exec"
            ),
            ns,
        )
        self.StorageService = ns["StorageService"]
        self._safe_tx = ns["_safe_tx"]

    # ── _safe_tx ──────────────────────────────────────────────────────────────

    @pytest.mark.parametrize("inp,expected", [
        ("T0001",    "T0001"),
        ("260006",   "260006"),
        ("26/27",    "26-27"),
        ("INV/2026", "INV-2026"),
        ("A\\B",     "A-B"),
        ("A B",      "A-B"),
        ("a//b",     "a-b"),
    ])
    def test_safe_tx_sanitizes(self, inp, expected):
        assert self._safe_tx(inp) == expected

    def test_safe_tx_empty_returns_unknown(self):
        assert self._safe_tx("") == "UNKNOWN"
        assert self._safe_tx("   ") == "UNKNOWN"

    # ── build_output_path ─────────────────────────────────────────────────────

    def test_path_no_slash_in_tx_no(self):
        with tempfile.TemporaryDirectory() as tmp:
            svc = self.StorageService(tmp)
            path = svc.build_output_path("T0001", "INV-202601-0001", "ar", "document")
            assert "T0001" in str(path)
            assert path.suffix == ".pdf"

    def test_path_slash_tx_sanitized(self):
        with tempfile.TemporaryDirectory() as tmp:
            svc = self.StorageService(tmp)
            path = svc.build_output_path("26/27", "INV-202602-0002", "ar", "document")
            parts = path.parts
            # No part should be "26" or "27" separately
            assert "26-27" in parts
            assert "26" not in parts[:-3]  # not in the folder hierarchy

    def test_path_creates_directory(self):
        with tempfile.TemporaryDirectory() as tmp:
            svc = self.StorageService(tmp)
            path = svc.build_output_path("T0005", "PL-202601-0001", "en", "document")
            assert path.parent.exists()

    def test_path_from_doc_no_extracts_date(self):
        with tempfile.TemporaryDirectory() as tmp:
            svc = self.StorageService(tmp)
            path = svc.build_output_path_from_doc_no(
                "T0003", "INV-202510-0007", "ar", "document"
            )
            assert "2025" in str(path)
            assert "10" in str(path)


# ══════════════════════════════════════════════════════════════════════════════
# NumberingService — pure helpers
# ══════════════════════════════════════════════════════════════════════════════

class TestNumberingServicePure:

    @pytest.fixture(autouse=True)
    def _import(self):
        import sys
        sys.path.insert(0, str(Path(__file__).parent.parent))
        from services.numbering_service import NumberingService
        self.NS = NumberingService

    def test_is_numeric_transaction_numeric(self):
        assert self.NS.is_numeric_transaction("T0001") is True
        assert self.NS.is_numeric_transaction("260006") is True

    def test_is_numeric_transaction_slash(self):
        assert self.NS.is_numeric_transaction("26/27") is False

    def test_extract_numeric_part_with_prefix(self):
        result = self.NS.extract_numeric_part("T0007")
        assert result == 7

    def test_extract_numeric_part_no_prefix(self):
        result = self.NS.extract_numeric_part("260006")
        assert result == 260006

    def test_extract_numeric_part_slash_strips_to_digits(self):
        """extract_numeric_part strips non-digits — '26/27' → 2627 (digits concatenated)."""
        result = self.NS.extract_numeric_part("26/27")
        # The function strips all non-digits, so 26/27 → 2627
        assert result == 2627

    def test_extract_numeric_part_pure_text_returns_none(self):
        result = self.NS.extract_numeric_part("ALPHA")
        assert result is None

    def test_format_transaction_number(self):
        assert self.NS.format_transaction_number(7, "T") == "T7"
        assert self.NS.format_transaction_number(260006, "") == "260006"


# ══════════════════════════════════════════════════════════════════════════════
# Documents DB — doc_groups and documents (requires in-memory DB)
# ══════════════════════════════════════════════════════════════════════════════

class TestDocumentsDB:

    def test_doc_group_created_with_transaction(
        self, db_session, make_transaction
    ):
        """Creating a transaction allows doc_groups to be linked."""
        from sqlalchemy import text
        trx = make_transaction(transaction_no="TDOC001")

        # Manually insert a doc_group (simulating what persist_document does)
        db_session.execute(text("""
            INSERT INTO doc_groups (transaction_id, doc_no, year, month, seq, created_at)
            VALUES (:tid, 'INV-202602-0001', 2026, 2, 1, datetime('now'))
        """), {"tid": trx.id})
        db_session.flush()

        row = db_session.execute(
            text("SELECT doc_no FROM doc_groups WHERE transaction_id=:tid"),
            {"tid": trx.id}
        ).fetchone()
        assert row is not None
        assert row[0] == "INV-202602-0001"

    def test_document_linked_to_doc_group(
        self, db_session, make_transaction
    ):
        """Documents can be linked to doc_groups."""
        from sqlalchemy import text

        trx = make_transaction(transaction_no="TDOC002")

        # Insert doc_group
        db_session.execute(text("""
            INSERT INTO doc_groups (transaction_id, doc_no, year, month, seq, created_at)
            VALUES (:tid, 'INV-202602-0002', 2026, 2, 1, datetime('now'))
        """), {"tid": trx.id})
        db_session.flush()

        group_id = db_session.execute(
            text("SELECT id FROM doc_groups WHERE transaction_id=:tid"), {"tid": trx.id}
        ).fetchone()[0]

        # Insert document_type needed for FK
        db_session.execute(text("""
            INSERT OR IGNORE INTO document_types (id, code, is_active, sort_order)
            VALUES (1, 'INV_COM', 1, 1)
        """))
        db_session.flush()

        db_session.execute(text("""
            INSERT INTO documents (group_id, document_type_id, language, status, created_at)
            VALUES (:gid, 1, 'ar', 'generated', datetime('now'))
        """), {"gid": group_id})
        db_session.flush()

        doc = db_session.execute(
            text("SELECT language, status FROM documents WHERE group_id=:gid"),
            {"gid": group_id}
        ).fetchone()
        assert doc is not None
        assert doc[0] == "ar"
        assert doc[1] == "generated"

    def test_cascade_delete_doc_group_on_transaction_delete(
        self, db_session, make_transaction
    ):
        """Deleting a transaction should cascade to doc_groups (FK CASCADE)."""
        from sqlalchemy import text

        trx = make_transaction(transaction_no="TDOC003")
        trx_id = trx.id

        db_session.execute(text("""
            INSERT INTO doc_groups (transaction_id, doc_no, year, month, seq, created_at)
            VALUES (:tid, 'INV-202602-0003', 2026, 2, 1, datetime('now'))
        """), {"tid": trx_id})
        db_session.flush()

        # Verify it exists
        before = db_session.execute(
            text("SELECT COUNT(*) FROM doc_groups WHERE transaction_id=:tid"),
            {"tid": trx_id}
        ).scalar()
        assert before == 1

        # Delete transaction
        db_session.delete(trx)
        db_session.flush()

        # Check cascade
        after = db_session.execute(
            text("SELECT COUNT(*) FROM doc_groups WHERE transaction_id=:tid"),
            {"tid": trx_id}
        ).scalar()
        assert after == 0, "doc_groups should be deleted when transaction is deleted (CASCADE)"