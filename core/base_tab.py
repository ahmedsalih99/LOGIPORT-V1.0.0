"""
core/base_tab.py
================
Base class for all data table tabs in LOGIPORT.

Architecture:
  - setup_ui()       → بناء الـ layout الأساسي
  - setup_table()    → إعدادات الجدول
  - setup_shortcuts()→ اختصارات لوحة المفاتيح
  - setup_signals()  → ربط الإشارات
  - check_permissions() → إظهار/إخفاء الأزرار حسب الصلاحيات

Column format: {"label": "i18n_key", "key": "data_key", "align": Qt.AlignCenter}
"""

import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from PySide6.QtCore import Qt, QModelIndex, Signal, QTimer, QEvent, QObject
from PySide6.QtGui import QKeySequence, QShortcut, QGuiApplication, QFont
from PySide6.QtWidgets import (
    QWidget, QTableWidget, QAbstractItemView, QMenu, QVBoxLayout, QHBoxLayout,
    QPushButton, QSpacerItem, QSizePolicy, QLabel, QComboBox, QLineEdit,
    QFileDialog, QMessageBox, QTableWidgetItem, QHeaderView, QCheckBox,
    QAbstractSpinBox, QApplication, QFrame,
)

from core.settings_manager import SettingsManager
from core.translator import TranslationManager
from core.permissions import is_admin as _is_admin, has_perm as _has_perm, has_any_perm

logger = logging.getLogger(__name__)

# ── ثوابت ارتفاع الصفوف الافتراضية ─────────────────────────────────────────
# القيمة الافتراضية لارتفاع الصف — يمكن للتابات/الديالوغات تجاوزها
TABLE_ROW_HEIGHT_DEFAULT = 46   # px
TABLE_ROW_HEIGHT_MIN     = 32   # px الحد الأدنى


def _make_table_item_font(size_px: int | None = None, bold: bool = True) -> QFont:
    """
    ينشئ QFont لخلايا الجداول يتناسق مع إعدادات التطبيق الحالية.
    يقرأ font_family و font_size من ThemeManager إن لم تُحدَّد.
    bold=True دائماً لنص خلايا الجداول (معيار التطبيق).
    """
    try:
        from core.theme_manager import ThemeManager
        tm = ThemeManager.get_instance()
        family = tm.get_current_font_family()
        if size_px is None:
            size_px = tm.get_current_font_size()
    except Exception:
        family = "Tajawal"
        if size_px is None:
            size_px = 12
    f = QFont(family, size_px)
    f.setBold(bold)
    return f


# فونت مشترك لخلايا الجداول — يُعاد بناؤه عند تغيير الثيم (انظر _refresh_table_font)
_BOLD_ITEM_FONT = _make_table_item_font()


# ── منع scroll على QComboBox/QSpinBox ────────────────────────────────────────
class _NoWheelOnInputs(QObject):
    """
    EventFilter يمنع QComboBox و QSpinBox من تغيير قيمتها بعجلة الفأرة.
    يُنصَّب مرة واحدة على QApplication فيغطّي كل الـ widgets.
    """
    def eventFilter(self, obj: QObject, event: QEvent) -> bool:
        if event.type() == QEvent.Type.Wheel and isinstance(obj, (QComboBox, QAbstractSpinBox)):
            event.ignore()
            return True
        return super().eventFilter(obj, event)


_wheel_filter = _NoWheelOnInputs()




# ─────────────────────────────────────────────────────────────────────────────
class DateRangeBar(QWidget):
    """
    شريط فلترة بالتاريخ مشترك بين transactions / entries / audit_trail.

    الاستخدام في أي تاب:
        self.date_bar = DateRangeBar(self)
        self.date_bar.changed.connect(self.reload_data)
        self.layout.insertWidget(1, self.date_bar)

    قراءة القيم:
        date_from_str, date_to_str = self.date_bar.get_range()  # "yyyy-MM-dd"

    إضافة widgets إضافية (نوع / حالة ...):
        self.date_bar.add_widget(my_combo)
        self.date_bar.add_separator()

    إضافة label عدد النتائج:
        self.date_bar.set_count(n)
    """

    changed = Signal()   # يُصدر عند تغيير أي فلتر

    def __init__(self, parent=None, default_months: int = 3):
        super().__init__(parent)
        self.setObjectName("filter-bar")
        self._     = TranslationManager.get_instance().translate
        self._build(default_months)

    def _build(self, default_months: int):
        from PySide6.QtCore import QDate
        self._main_lay = QHBoxLayout(self)
        self._main_lay.setContentsMargins(8, 4, 8, 4)
        self._main_lay.setSpacing(6)

        # ── نطاق التاريخ ─────────────────────────────────────────────
        date_grp = QWidget()
        date_grp.setObjectName("filter-group")
        date_lay = QHBoxLayout(date_grp)
        date_lay.setContentsMargins(8, 3, 8, 3)
        date_lay.setSpacing(5)

        from PySide6.QtWidgets import QDateEdit
        icon = QLabel("📅")
        icon.setFixedWidth(18)
        date_lay.addWidget(icon)

        self._date_from = QDateEdit()
        self._date_from.setObjectName("form-input")
        self._date_from.setCalendarPopup(True)
        self._date_from.setDisplayFormat("yyyy-MM-dd")
        self._date_from.setDate(QDate.currentDate().addMonths(-default_months))
        self._date_from.setFixedWidth(104)
        self._date_from.dateChanged.connect(self._on_changed)
        date_lay.addWidget(self._date_from)

        arr = QLabel("→")
        arr.setObjectName("text-muted")
        arr.setStyleSheet("font-weight:600; padding:0 2px;")
        date_lay.addWidget(arr)

        self._date_to = QDateEdit()
        self._date_to.setObjectName("form-input")
        self._date_to.setCalendarPopup(True)
        self._date_to.setDisplayFormat("yyyy-MM-dd")
        self._date_to.setDate(QDate.currentDate())
        self._date_to.setFixedWidth(104)
        self._date_to.dateChanged.connect(self._on_changed)
        date_lay.addWidget(self._date_to)

        self._main_lay.addWidget(date_grp)

        # ── أزرار Preset ──────────────────────────────────────────────
        self._btn_today = self._preset_btn("today",      self._set_today)
        self._btn_week  = self._preset_btn("this_week",  self._set_week)
        self._btn_month = self._preset_btn("this_month", self._set_month)
        self._btn_clear = self._clear_btn()

        for btn in (self._btn_today, self._btn_week, self._btn_month, self._btn_clear):
            self._main_lay.addWidget(btn)

        # ── فاصل + منطقة إضافية (type/status combos ...) ─────────────
        self._extra_sep = self._sep()
        self._extra_sep.setVisible(False)
        self._main_lay.addWidget(self._extra_sep)

        # stretch
        self._main_lay.addStretch(1)

        # label العدد
        self._count_lbl = QLabel()
        self._count_lbl.setObjectName("filter-count-lbl")
        self._count_lbl.hide()
        self._main_lay.addWidget(self._count_lbl)

    # ── helpers بناء ─────────────────────────────────────────────────

    def _preset_btn(self, key: str, slot) -> QPushButton:
        btn = QPushButton(self._(key))
        btn.setObjectName("filter-preset-btn")
        btn.setFixedHeight(30)
        btn.setCursor(Qt.PointingHandCursor)
        btn.clicked.connect(slot)
        return btn

    def _clear_btn(self) -> QPushButton:
        btn = QPushButton("✖")
        btn.setObjectName("filter-clear-btn")
        btn.setFixedSize(30, 30)
        btn.setCursor(Qt.PointingHandCursor)
        btn.setToolTip(self._("clear"))
        btn.clicked.connect(self._set_clear)
        return btn

    def _sep(self) -> QFrame:
        sep = QFrame()
        sep.setObjectName("filter-sep")
        sep.setFrameShape(QFrame.VLine)
        sep.setFixedHeight(24)
        return sep

    # ── preset slots ─────────────────────────────────────────────────

    def _on_changed(self):
        self.changed.emit()

    def _set_today(self):
        from PySide6.QtCore import QDate
        today = QDate.currentDate()
        self._date_from.dateChanged.disconnect(self._on_changed)
        self._date_from.setDate(today)
        self._date_from.dateChanged.connect(self._on_changed)
        self._date_to.setDate(today)

    def _set_week(self):
        from PySide6.QtCore import QDate
        today = QDate.currentDate()
        self._date_from.dateChanged.disconnect(self._on_changed)
        self._date_from.setDate(today.addDays(-today.dayOfWeek() + 1))
        self._date_from.dateChanged.connect(self._on_changed)
        self._date_to.setDate(today)

    def _set_month(self):
        from PySide6.QtCore import QDate
        today = QDate.currentDate()
        self._date_from.dateChanged.disconnect(self._on_changed)
        self._date_from.setDate(QDate(today.year(), today.month(), 1))
        self._date_from.dateChanged.connect(self._on_changed)
        self._date_to.setDate(today)

    def _set_clear(self):
        from PySide6.QtCore import QDate
        self._date_from.dateChanged.disconnect(self._on_changed)
        self._date_from.setDate(QDate.currentDate().addMonths(-3))
        self._date_from.dateChanged.connect(self._on_changed)
        self._date_to.setDate(QDate.currentDate())

    # ── Public API ────────────────────────────────────────────────────

    def get_range(self) -> tuple:
        """يُرجع (date_from_str, date_to_str) بصيغة yyyy-MM-dd."""
        return (
            self._date_from.date().toString("yyyy-MM-dd"),
            self._date_to.date().toString("yyyy-MM-dd"),
        )

    def add_separator(self):
        """يضيف فاصل عمودي قبل الـ stretch."""
        sep = self._sep()
        self._extra_sep.setVisible(True)
        idx = self._main_lay.indexOf(self._extra_sep)
        self._main_lay.insertWidget(idx + 1, sep)

    def add_widget(self, widget: QWidget):
        """يضيف widget (combo/label...) قبل الـ stretch — مع فاصل تلقائي أول مرة."""
        if not self._extra_sep.isVisible():
            self._extra_sep.setVisible(True)
        idx = self._main_lay.indexOf(self._count_lbl)
        self._main_lay.insertWidget(idx, widget)

    def set_count(self, n: int):
        """يحدّث label عدد النتائج."""
        _ = TranslationManager.get_instance().translate
        self._count_lbl.setText(f"{n}  {_('total_rows')}")
        self._count_lbl.show()

    def retranslate(self):
        _ = TranslationManager.get_instance().translate
        self._btn_today.setText(_("today"))
        self._btn_week.setText(_("this_week"))
        self._btn_month.setText(_("this_month"))
        self._btn_clear.setToolTip(_("clear"))


# ─────────────────────────────────────────────────────────────────────────────
class BaseTab(QWidget):
    """
    الكلاس الأساسي لكل تابات الجداول.

    يوفر:
      - شريط أدوات موحد (بحث + أزرار إضافة/تصدير/تحديث)
      - جدول بيانات مع pagination
      - تصدير Excel مع تنسيق احترافي
      - دعم أعمدة الأدمن (toggle)
      - اختصارات لوحة المفاتيح
      - حماية scroll على ComboBox/SpinBox

    ما يوفره عمداً (متروك للتابات الفرعية):
      - طباعة (print) — غير مستخدمة في أي تاب حالياً
      - استيراد Excel (import) — غير مُكتمل ولا يُسجّل في DB
      - filter_box (ترتيب) — تابات مثل transactions تُخفيه وتستخدم combo خاص بها
    """

    row_double_clicked = Signal(int)
    request_edit       = Signal(int)
    request_delete     = Signal(list)
    request_view       = Signal(int)

    required_permissions: dict = {
        "add":     None,
        "export":  None,
        "refresh": None,
        "edit":    None,
        "delete":  None,
        "view":    None,
    }

    def __init__(self, title=None, parent=None, user=None):
        super().__init__(parent)
        self.title        = title
        self.settings     = SettingsManager.get_instance()
        self._            = TranslationManager.get_instance().translate
        self.user         = user or self.settings.get("user") or {}
        self.current_user = self.user
        self.is_admin     = _is_admin(self.user)

        # ── pagination state ─────────────────────────────────────────────
        self.rows_per_page = 20
        self.current_page  = 1
        self.total_rows    = 0
        self.total_pages   = 1
        self.data: list    = []

        # ── columns ──────────────────────────────────────────────────────
        self.columns: list       = []
        self._base_columns: list = []
        self._admin_columns: list= []

        # ── ارتفاع صفوف الجدول ───────────────────────────────────────────
        # التابات الفرعية تقدر تغير هذه القيمة قبل _setup_table
        # أو بعده باستخدام set_row_height()
        self.table_row_height: int = TABLE_ROW_HEIGHT_DEFAULT

        # ── build ────────────────────────────────────────────────────────
        self._setup_ui()
        self._setup_table()
        self._setup_pagination_controls()
        self._setup_shortcuts()
        self._setup_signals()

        TranslationManager.get_instance().language_changed.connect(self.retranslate_ui)

        # حجب ScrollWheel على كل QComboBox و QSpinBox في الـ tab
        try:
            from ui.utils.wheel_blocker import block_wheel_in
            block_wheel_in(self)
        except Exception:
            pass

    # ─────────────────────────────────────────────────────────────────────
    # UI BUILD
    # ─────────────────────────────────────────────────────────────────────

    def _setup_ui(self):
        self._layout = QVBoxLayout(self)
        self._layout.setContentsMargins(6, 6, 6, 6)
        self._layout.setSpacing(4)

        # ── شريط الأدوات ──────────────────────────────────────────────
        self.top_bar = QHBoxLayout()
        self.top_bar.setSpacing(5)

        self.search_bar = QLineEdit()
        self.search_bar.setPlaceholderText(self._("search") + "...")
        self.search_bar.setObjectName("search-field")
        self.search_bar.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.top_bar.addWidget(self.search_bar)

        # spacer
        self.top_bar.addSpacerItem(QSpacerItem(8, 0, QSizePolicy.Expanding, QSizePolicy.Minimum))

        # أعمدة الأدمن — checkbox مرئي للأدمن فقط
        self.chk_admin_cols = QCheckBox(self._("show_admin_columns"))
        self.chk_admin_cols.setVisible(self.is_admin)
        self.chk_admin_cols.setChecked(False)
        self.chk_admin_cols.stateChanged.connect(self._apply_columns_for_current_role)
        self.top_bar.addWidget(self.chk_admin_cols)

        # ── أزرار الأدوات ──────────────────────────────────────────────
        self.btn_add     = self._toolbar_btn("add",              "add")
        self.btn_export  = self._toolbar_btn("export to excel",  "export")
        self.btn_refresh = self._toolbar_btn("refresh",          "refresh")

        for btn in (self.btn_add, self.btn_export, self.btn_refresh):
            self.top_bar.addWidget(btn)

        self._layout.addLayout(self.top_bar)

        # ── الجدول ────────────────────────────────────────────────────
        self.table = QTableWidget(0, 0, self)
        self.table.setObjectName("data-table")
        self._layout.addWidget(self.table)

        # ── شريط الحالة (عدد الصفوف + ترتيب) ─────────────────────────
        status_bar = QHBoxLayout()
        status_bar.setContentsMargins(4, 0, 4, 0)

        self._lbl_count = QLabel()
        self._lbl_count.setObjectName("status-count-lbl")
        self._lbl_sort  = QLabel()
        self._lbl_sort.setObjectName("status-sort-lbl")

        status_bar.addWidget(self._lbl_count)
        status_bar.addStretch(1)
        status_bar.addWidget(self._lbl_sort)
        self._layout.addLayout(status_bar)

        # ── empty state overlay ────────────────────────────────────────
        self._empty_widget = QWidget(self.table)
        self._empty_widget.setObjectName("empty-state")
        _ev = QVBoxLayout(self._empty_widget)
        _ev.setAlignment(Qt.AlignCenter)
        self._lbl_empty_icon = QLabel("📋")
        self._lbl_empty_icon.setAlignment(Qt.AlignCenter)
        self._lbl_empty_icon.setStyleSheet("font-size: 36px;")
        self._lbl_empty_text = QLabel(self._("no_data_available"))
        self._lbl_empty_text.setAlignment(Qt.AlignCenter)
        self._lbl_empty_text.setObjectName("empty-state-text")
        _ev.addWidget(self._lbl_empty_icon)
        _ev.addWidget(self._lbl_empty_text)
        self._empty_widget.hide()

        # ── pagination ────────────────────────────────────────────────
        self.pagination_bar = QHBoxLayout()

        self.btn_prev = QPushButton("◀")
        self.btn_prev.setObjectName("pagination-btn")
        self.btn_prev.setFixedWidth(32)

        self.btn_next = QPushButton("▶")
        self.btn_next.setObjectName("pagination-btn")
        self.btn_next.setFixedWidth(32)

        self.lbl_pagination = QLabel()
        self.lbl_pagination.setAlignment(Qt.AlignCenter)
        self.lbl_pagination.setObjectName("pagination-lbl")

        self.cmb_rows_per_page = QComboBox()
        self.cmb_rows_per_page.addItems(["10", "20", "50", "100"])
        self.cmb_rows_per_page.setCurrentText(str(self.rows_per_page))
        self.cmb_rows_per_page.setFixedWidth(64)

        self._lbl_rows_per_page = QLabel(self._("rows_per_page"))

        self.pagination_bar.addStretch(1)
        self.pagination_bar.addWidget(self.btn_prev)
        self.pagination_bar.addWidget(self.lbl_pagination)
        self.pagination_bar.addWidget(self.btn_next)
        self.pagination_bar.addStretch(1)
        self.pagination_bar.addWidget(self._lbl_rows_per_page)
        self.pagination_bar.addWidget(self.cmb_rows_per_page)
        self._layout.addLayout(self.pagination_bar)

    def _toolbar_btn(self, label_key: str, obj_name: str) -> QPushButton:
        btn = QPushButton(self._(label_key))
        btn.setObjectName("action-btn")
        btn.setMinimumWidth(72)
        btn.setMaximumWidth(130)
        return btn

    def _setup_table(self):
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.setSortingEnabled(True)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.horizontalHeader().setSectionsMovable(False)
        self.table.horizontalHeader().setStretchLastSection(True)

        # ── ارتفاع الصفوف: قيمة افتراضية + إمكانية تعديل يدوي ──────────
        self._apply_row_height(self.table_row_height)

        # ── عرض الأعمدة: Interactive (قابل للسحب) ─────────────────────
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)

        # ── رأس الجدول: ارتفاع + فونت Bold ───────────────────────────
        self._apply_header_style()

        # ── signals ─────────────────────────────────────────────────────
        self.table.customContextMenuRequested.connect(self._show_context_menu)
        self.table.doubleClicked.connect(self._on_row_double_clicked)

        # ── تحديث الفونت عند تغيير الثيم ──────────────────────────────
        try:
            from core.theme_manager import ThemeManager
            ThemeManager.get_instance().theme_changed.connect(self._on_theme_changed)
        except Exception:
            pass

    def _apply_row_height(self, height: int):
        """يطبّق ارتفاع الصفوف — يُستدعى من _setup_table وعند تغيير الثيم."""
        height = max(height, TABLE_ROW_HEIGHT_MIN)
        self.table.verticalHeader().setDefaultSectionSize(height)
        self.table.verticalHeader().setMinimumSectionSize(TABLE_ROW_HEIGHT_MIN)

    def _apply_header_style(self):
        """يطبّق ارتفاع رأس الجدول وفونته Bold — يتناسق مع font_size الحالي."""
        global _BOLD_ITEM_FONT
        _BOLD_ITEM_FONT = _make_table_item_font()

        # ارتفاع رأس الجدول: font_size × 3 + padding (بحد أدنى 40)
        try:
            from core.theme_manager import ThemeManager
            fs = ThemeManager.get_instance().get_current_font_size()
        except Exception:
            fs = 12
        hdr_height = max(40, fs * 3 + 8)
        self.table.horizontalHeader().setMinimumHeight(hdr_height)

        hdr_font = _make_table_item_font(bold=True)
        self.table.horizontalHeader().setFont(hdr_font)

    def _on_theme_changed(self, *_):
        """عند تغيير الثيم: يعيد بناء الفونت ويعيد تطبيق الأحجام."""
        self._apply_header_style()
        self._apply_row_height(self.table_row_height)
        # إعادة تطبيق الفونت على الصفوف الموجودة
        global _BOLD_ITEM_FONT
        _BOLD_ITEM_FONT = _make_table_item_font()
        for row in range(self.table.rowCount()):
            for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item:
                    item.setFont(_BOLD_ITEM_FONT)

    def _setup_pagination_controls(self):
        self.cmb_rows_per_page.currentTextChanged.connect(self._on_rows_per_page_changed)
        self.btn_prev.clicked.connect(self.go_to_prev_page)
        self.btn_next.clicked.connect(self.go_to_next_page)
        self._update_pagination_label()

    def _setup_shortcuts(self):
        QShortcut(QKeySequence("Ctrl+N"),          self, self.add_new_item)
        QShortcut(QKeySequence("Ctrl+E"),          self, self.edit_selected_item)
        QShortcut(QKeySequence("Ctrl+A"),          self, self.select_all_items)
        QShortcut(QKeySequence("Ctrl+D"),          self, self.clear_selection)
        QShortcut(QKeySequence("PageUp"),          self, self.go_to_prev_page)
        QShortcut(QKeySequence("PageDown"),        self, self.go_to_next_page)
        QShortcut(QKeySequence("Ctrl+Shift+E"),    self, self.export_table_to_excel)
        QShortcut(QKeySequence("Ctrl+R"),          self, self.refresh_data)
        QShortcut(QKeySequence("Ctrl+F"),          self, self.focus_search)
        QShortcut(QKeySequence("Ctrl+C"),          self, self.copy_selected)
        QShortcut(QKeySequence("Delete"),          self,
                  lambda: self.delete_selected_items() if self._table_has_focus() else None)
        QShortcut(QKeySequence("Return"),          self,
                  lambda: self.view_selected_item() if self._table_has_focus() else None)
        QShortcut(QKeySequence("Escape"),          self,
                  lambda: self.clear_selection() if self._table_has_focus() else None)

    def _setup_signals(self):
        self.btn_add.clicked.connect(self.add_new_item)
        self.btn_export.clicked.connect(self.export_table_to_excel)
        self.btn_refresh.clicked.connect(self.refresh_data)

        # debounce بحث — 350ms بعد آخر حرف
        self._search_timer = QTimer(self)
        self._search_timer.setSingleShot(True)
        self._search_timer.setInterval(350)
        self._search_timer.timeout.connect(self._on_search_changed)
        self.search_bar.textChanged.connect(lambda _: self._search_timer.start())

    # ─────────────────────────────────────────────────────────────────────
    # PUBLIC SETUP API  (تُستدعى من التابات الفرعية في __init__)
    # ─────────────────────────────────────────────────────────────────────

    def setup_filters(self):
        """Override في التابات الفرعية لإضافة widgets فلترة إضافية."""
        pass

    def set_columns(self, columns: list):
        """تعيين الأعمدة مباشرة."""
        self.columns = columns or []
        self.table.setColumnCount(len(self.columns))
        self.table.setHorizontalHeaderLabels([
            self._(c.get("label", "")) if c.get("label") else ""
            for c in self.columns
        ])

    def set_columns_for_role(self, base_columns: list, admin_columns: list = None):
        """تعيين الأعمدة الأساسية وأعمدة الأدمن الاختيارية."""
        self._base_columns  = base_columns or []
        self._admin_columns = admin_columns or []
        self._apply_columns_for_current_role()

    def check_permissions(self):
        """يُظهر/يُخفي الأزرار حسب required_permissions."""
        mapping = [
            (self.btn_add,     "add"),
            (self.btn_export,  "export"),
            (self.btn_refresh, "refresh"),
        ]
        for btn, key in mapping:
            perm = self.required_permissions.get(key)
            if perm:
                visible = (
                    has_any_perm(self.user, list(perm))
                    if isinstance(perm, (list, tuple, set))
                    else _has_perm(self.user, perm)
                )
                btn.setVisible(visible)
            else:
                btn.setVisible(True)

    def set_current_user(self, user):
        self.current_user = user
        self.user         = user
        try:
            self.is_admin = _is_admin(user)
            self.chk_admin_cols.setVisible(self.is_admin)
        except Exception:
            pass
        try:
            self._apply_columns_for_current_role()
        except Exception:
            pass
        try:
            self.check_permissions()
        except Exception:
            pass

    # ─────────────────────────────────────────────────────────────────────
    # COLUMNS
    # ─────────────────────────────────────────────────────────────────────

    def _apply_columns_for_current_role(self):
        cols = list(self._base_columns)
        if self.is_admin and self.chk_admin_cols.isChecked():
            cols += self._admin_columns
        self.set_columns(cols)

    def _apply_admin_columns(self):
        """يُخفي أعمدة الأدمن (id, created_by…) للمستخدمين العاديين."""
        admin_keys = {"id", "created_by_name", "updated_by_name", "created_at", "updated_at"}
        for idx, col in enumerate(self.columns):
            if col.get("key") in admin_keys:
                try:
                    self.table.setColumnHidden(idx, not self.is_admin)
                except Exception:
                    pass

    # ─────────────────────────────────────────────────────────────────────
    # DATA DISPLAY
    # ─────────────────────────────────────────────────────────────────────

    def reload_data(self):
        """يُعاد تعريفه في التابات الفرعية لجلب البيانات وتحديث self.data."""
        self.display_data()

    def display_data(self):
        """عرض self.data في الجدول مع بحث + ترتيب + pagination."""
        rows = list(self.data) if self.data else []
        total_before = len(rows)
        searched = bool((self.search_bar.text() or "").strip())

        if not getattr(self, "_skip_base_search", False):
            rows = self._apply_base_search(rows)
        if not getattr(self, "_skip_base_sort", False):
            rows = self._apply_base_sort(rows)

        self.total_rows  = len(rows)
        self.total_pages = max(1, (self.total_rows + self.rows_per_page - 1) // self.rows_per_page)
        self.current_page = max(1, min(self.current_page, self.total_pages))
        start     = (self.current_page - 1) * self.rows_per_page
        page_rows = rows[start: start + self.rows_per_page]

        self._update_pagination_label()
        self._update_status_bar(len(rows), total_before)
        self._show_empty_state(len(rows) == 0, searched=searched)

        if not self.columns:
            self.table.setRowCount(0)
            return

        self.table.setSortingEnabled(False)
        self.table.setUpdatesEnabled(False)
        try:
            self.table.setRowCount(len(page_rows))
            for i, row in enumerate(page_rows):
                for j, col in enumerate(self.columns):
                    key  = col.get("key", "")
                    val  = row.get(key, "")
                    item = QTableWidgetItem(str(val) if val is not None else "")
                    item.setTextAlignment(col.get("align", Qt.AlignCenter))
                    item.setFont(_BOLD_ITEM_FONT)
                    self.table.setItem(i, j, item)
        finally:
            self.table.setUpdatesEnabled(True)
            self.table.setSortingEnabled(True)
        self._stretch_columns()

    def _display_with_actions(self, edit_perm: str, delete_perm: str):
        """
        عرض self.data مع أزرار Edit/Delete في عمود 'actions'.
        الشرط: يجب أن يعرّف التاب الفرعي:
          - _open_edit_dialog(obj)
          - _delete_single(obj)
        """
        can_edit     = _has_perm(self.current_user, edit_perm)   if edit_perm   else False
        can_delete   = _has_perm(self.current_user, delete_perm) if delete_perm else False
        show_actions = can_edit or can_delete

        rows = list(self.data) if self.data else []
        total_before = len(rows)
        searched = bool((self.search_bar.text() or "").strip())

        if not getattr(self, "_skip_base_search", False):
            rows = self._apply_base_search(rows)
        if not getattr(self, "_skip_base_sort", False):
            rows = self._apply_base_sort(rows)

        self.total_rows  = len(rows)
        self.total_pages = max(1, (self.total_rows + self.rows_per_page - 1) // self.rows_per_page)
        self.current_page = max(1, min(self.current_page, self.total_pages))
        start     = (self.current_page - 1) * self.rows_per_page
        page_rows = rows[start: start + self.rows_per_page]

        self._update_pagination_label()
        self._update_status_bar(len(rows), total_before)
        self._show_empty_state(len(rows) == 0, searched=searched)

        self.table.setSortingEnabled(False)
        self.table.setUpdatesEnabled(False)
        try:
            self.table.setRowCount(len(page_rows))
            for row_idx, row in enumerate(page_rows):
                for col_idx, col in enumerate(self.columns):
                    key = col.get("key", "")
                    if key == "actions":
                        if not show_actions:
                            continue
                        obj = row.get("actions")
                        self._set_action_cell(row_idx, col_idx, obj, can_edit, can_delete)
                    else:
                        val  = row.get(key, "")
                        item = QTableWidgetItem(str(val) if val is not None else "")
                        item.setTextAlignment(col.get("align", Qt.AlignCenter))
                        item.setFont(_BOLD_ITEM_FONT)
                        self.table.setItem(row_idx, col_idx, item)
        finally:
            self.table.setUpdatesEnabled(True)
            self.table.setSortingEnabled(True)

        # إخفاء عمود actions إذا لا صلاحية
        try:
            ai = next((i for i, c in enumerate(self.columns) if c.get("key") == "actions"), None)
            if ai is not None:
                self.table.setColumnHidden(ai, not show_actions)
        except Exception:
            pass

        self._apply_admin_columns()
        self._stretch_columns()

    def _set_action_cell(self, row_idx, col_idx, obj, can_edit, can_delete):
        """يبني cell الأزرار لعمود actions."""
        if can_edit and can_delete:
            w   = QWidget()
            lay = QHBoxLayout(w)
            lay.setContentsMargins(2, 2, 2, 2)
            lay.setSpacing(3)
            btn_e = QPushButton(self._("edit"))
            btn_e.setObjectName("table-edit")
            btn_e.clicked.connect(lambda _=False, o=obj: self._open_edit_dialog(o))
            btn_d = QPushButton(self._("delete"))
            btn_d.setObjectName("table-delete")
            btn_d.clicked.connect(lambda _=False, o=obj: self._delete_single(o))
            lay.addWidget(btn_e)
            lay.addWidget(btn_d)
            self.table.setCellWidget(row_idx, col_idx, w)
        elif can_edit:
            btn = QPushButton(self._("edit"))
            btn.setObjectName("table-edit")
            btn.clicked.connect(lambda _=False, o=obj: self._open_edit_dialog(o))
            self.table.setCellWidget(row_idx, col_idx, btn)
        elif can_delete:
            btn = QPushButton(self._("delete"))
            btn.setObjectName("table-delete")
            btn.clicked.connect(lambda _=False, o=obj: self._delete_single(o))
            self.table.setCellWidget(row_idx, col_idx, btn)

    def _stretch_columns(self):
        """يضبط عرض الأعمدة تلقائياً حسب المحتوى ثم يتيح للمستخدم التعديل."""
        if not self.table.columnCount():
            return
        hdr = self.table.horizontalHeader()
        # أولاً: اضبط كل عمود حسب محتواه (header + cells)
        hdr.setSectionResizeMode(QHeaderView.ResizeToContents)
        # ثانياً: حوّل للوضع Interactive حتى يبقى التعديل اليدوي ممكناً
        # نؤخّر التحويل حتى تنتهي Qt من حساب الأحجام
        QTimer.singleShot(0, lambda: hdr.setSectionResizeMode(QHeaderView.Interactive)
                          if not self.table.isHidden() else None)

    def set_row_height(self, height: int):
        """
        API عام يسمح للتاب الفرعي أو الديالوغ بتعيين ارتفاع مخصص للصفوف.
        مثال:  self.set_row_height(52)
        """
        self.table_row_height = max(height, TABLE_ROW_HEIGHT_MIN)
        self._apply_row_height(self.table_row_height)

    # ─────────────────────────────────────────────────────────────────────
    # SEARCH / SORT
    # ─────────────────────────────────────────────────────────────────────

    def _get_search_keys(self) -> list:
        skip = {"id", "actions", "created_at", "updated_at", "created_by_name", "updated_by_name"}
        return [c.get("key", "") for c in self.columns if c.get("key") and c.get("key") not in skip]

    def _apply_base_search(self, rows: list) -> list:
        q = (self.search_bar.text() or "").strip().casefold()
        if not q:
            return rows
        keys = self._get_search_keys()
        return [r for r in rows if any(q in str(r.get(k, "") or "").casefold() for k in keys)]

    def _apply_base_sort(self, rows: list) -> list:
        return rows  # التابات الفرعية ترتّب server-side أو تعيد تعريف هذه

    def _on_search_changed(self):
        self.current_page = 1
        self.reload_data()

    # ─────────────────────────────────────────────────────────────────────
    # PAGINATION
    # ─────────────────────────────────────────────────────────────────────

    def _update_pagination_label(self):
        self.lbl_pagination.setText(
            f"{self._('page')} {self.current_page} / {self.total_pages}"
            f"  ({self._('total_rows')}: {self.total_rows})"
        )

    # backward-compat alias
    def update_pagination_label(self):
        self._update_pagination_label()

    def _on_rows_per_page_changed(self, value: str):
        self.rows_per_page = int(value)
        self.current_page  = 1
        self.reload_data()

    # backward-compat alias
    def on_rows_per_page_changed(self, value):
        self._on_rows_per_page_changed(value)

    def go_to_prev_page(self):
        if self.current_page > 1:
            self.current_page -= 1
            self.reload_data()

    def go_to_next_page(self):
        if self.current_page < self.total_pages:
            self.current_page += 1
            self.reload_data()

    # ─────────────────────────────────────────────────────────────────────
    # STATUS BAR / EMPTY STATE
    # ─────────────────────────────────────────────────────────────────────

    def _update_status_bar(self, displayed: int, total: int):
        q = (self.search_bar.text() or "").strip()
        if q and displayed < total:
            self._lbl_count.setText(f"🔍 {displayed} {self._('of')} {total} {self._('total_rows')}")
        else:
            self._lbl_count.setText(f"{displayed} {self._('total_rows')}")
        self._lbl_sort.setText("")

    def _show_empty_state(self, empty: bool, searched: bool = False):
        if empty:
            self._lbl_empty_icon.setText("🔍" if searched else "📋")
            self._lbl_empty_text.setText(
                self._("no_search_results") if searched else self._("no_data_available")
            )
            self._empty_widget.setGeometry(
                0, 44, self.table.width(), max(self.table.height() - 44, 100)
            )
            self._empty_widget.show()
            self._empty_widget.raise_()
        else:
            self._empty_widget.hide()

    # ─────────────────────────────────────────────────────────────────────
    # SELECT / NAVIGATE
    # ─────────────────────────────────────────────────────────────────────

    def select_record_by_id(self, record_id: int):
        """يبحث عن سجل بالـ ID ويحدده في الجدول."""
        if not self.data:
            self.reload_data()
        for row_idx, row in enumerate(self.data):
            if row.get("id") is not None and int(row["id"]) == int(record_id):
                self.table.setCurrentCell(row_idx, 0)
                self.table.scrollTo(self.table.model().index(row_idx, 0), self.table.PositionAtCenter)
                self.table.selectRow(row_idx)
                return
        try:
            self._reset_filters()
            self.reload_data()
            for row_idx, row in enumerate(self.data):
                if row.get("id") is not None and int(row["id"]) == int(record_id):
                    self.table.setCurrentCell(row_idx, 0)
                    self.table.scrollTo(self.table.model().index(row_idx, 0), self.table.PositionAtCenter)
                    self.table.selectRow(row_idx)
                    return
        except Exception:
            pass

    def _reset_filters(self):
        if hasattr(self, "search_bar"):
            self.search_bar.clear()

    def get_selected_rows(self) -> list:
        return [idx.row() for idx in self.table.selectionModel().selectedRows()]

    def _table_has_focus(self) -> bool:
        focused = QApplication.focusWidget()
        if focused is None:
            return False
        return not isinstance(focused, (QLineEdit, QComboBox))

    # ─────────────────────────────────────────────────────────────────────
    # CONTEXT MENU
    # ─────────────────────────────────────────────────────────────────────

    def _show_context_menu(self, pos):
        rows  = self.get_selected_rows()
        count = len(rows)
        menu  = QMenu(self)

        if count > 1:
            title = menu.addAction(f"✓  {count}  {self._('selected_rows_count')}")
            title.setEnabled(False)
            menu.addSeparator()

        act_view   = menu.addAction(self._("view"))
        act_edit   = menu.addAction(self._("edit"))
        act_view.setEnabled(count == 1)
        act_edit.setEnabled(count == 1)
        menu.addSeparator()

        act_copy = menu.addAction(self._("copy"))
        act_copy.setEnabled(count >= 1)
        menu.addSeparator()

        label_del  = f"🗑  {self._('delete')}  ({count})" if count > 1 else self._("delete")
        act_delete = menu.addAction(label_del)
        act_delete.setEnabled(count >= 1)

        # أي إجراءات إضافية من التاب الفرعي
        extra = self.get_extra_context_actions(menu)

        action = menu.exec(self.table.mapToGlobal(pos))
        if action is None:
            return
        if action == act_view   and rows: self.request_view.emit(rows[0])
        elif action == act_edit and rows: self.request_edit.emit(rows[0])
        elif action == act_copy and rows: self.copy_selected()
        elif action == act_delete and rows: self.request_delete.emit(rows)
        elif extra:
            for act, cb in extra:
                if action == act:
                    cb(rows)
                    break

    def get_extra_context_actions(self, menu) -> list:
        """Override في التابات الفرعية لإضافة عناصر للقائمة السياقية."""
        return []

    # ─────────────────────────────────────────────────────────────────────
    # EVENT HANDLERS
    # ─────────────────────────────────────────────────────────────────────

    def _on_row_double_clicked(self, index: QModelIndex):
        self.row_double_clicked.emit(index.row())
        self.request_view.emit(index.row())

    # backward-compat — بعض التابات تعيد تعريف on_row_double_clicked
    def on_row_double_clicked(self, index):
        if isinstance(index, QModelIndex):
            self._on_row_double_clicked(index)
        else:
            self.row_double_clicked.emit(int(index))

    # ─────────────────────────────────────────────────────────────────────
    # OVERRIDEABLE ACTIONS
    # ─────────────────────────────────────────────────────────────────────

    def add_new_item(self):        pass
    def edit_selected_item(self):  pass

    def delete_selected_items(self):
        rows = self.get_selected_rows()
        if rows:
            self.request_delete.emit(rows)

    def view_selected_item(self):
        rows = self.get_selected_rows()
        if rows:
            self.request_view.emit(rows[0])

    def select_all_items(self):   self.table.selectAll()
    def clear_selection(self):    self.table.clearSelection()
    def focus_search(self):       self.search_bar.setFocus()

    def refresh_data(self):
        self.reload_data()

    def copy_selected(self):
        """Ctrl+C — ينسخ خلايا الجدول المحددة بصيغة TSV (متوافقة مع Excel)."""
        selected = self.table.selectedItems()
        if not selected:
            return
        rows_data: dict = {}
        for item in selected:
            rows_data.setdefault(item.row(), {})[item.column()] = item.text()
        lines = ["\t".join(rows_data[r][c] for c in sorted(rows_data[r])) for r in sorted(rows_data)]
        QGuiApplication.clipboard().setText("\n".join(lines))

    # ─────────────────────────────────────────────────────────────────────
    # EXPORT TO EXCEL
    # ─────────────────────────────────────────────────────────────────────

    def export_table_to_excel(self):
        """تصدير البيانات المرئية في الجدول إلى Excel مع تنسيق احترافي."""
        default_name = f"{self.title or 'data'}.xlsx"
        path, _ = QFileDialog.getSaveFileName(
            self, self._("export to excel"), default_name, "Excel Files (*.xlsx)"
        )
        if not path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = (self.title or "Data")[:31]

            # الأعمدة المرئية (تستثني actions)
            col_indices = [
                i for i in range(self.table.columnCount())
                if not (i < len(self.columns) and self.columns[i].get("key") == "actions")
                and not self.table.isColumnHidden(i)
            ]
            headers = []
            for i in col_indices:
                h = self.table.horizontalHeaderItem(i)
                headers.append(h.text() if h else "")

            ws.append(headers)

            # تنسيق الهيدر
            hdr_font  = Font(bold=True, color="FFFFFF", size=11)
            hdr_fill  = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
            hdr_align = Alignment(horizontal="center", vertical="center")
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font      = hdr_font
                cell.fill      = hdr_fill
                cell.alignment = hdr_align

            # البيانات
            alt_fill  = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
            data_align = Alignment(horizontal="center", vertical="center")
            for row in range(self.table.rowCount()):
                values = []
                for ci in col_indices:
                    item = self.table.item(row, ci)
                    values.append(item.text() if item else "")
                ws.append(values)
                if row % 2 == 0:
                    for col_num in range(1, len(headers) + 1):
                        cell = ws.cell(row=row + 2, column=col_num)
                        cell.fill      = alt_fill
                        cell.alignment = data_align

            # عرض الأعمدة تلقائياً
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 3, 12), 50)

            ws.freeze_panes = "A2"
            wb.save(path)
            QMessageBox.information(self, self._("export to excel"),
                                    self._("Data exported successfully!") + f"\n\n{path}")
        except Exception as e:
            logger.exception(f"Export failed: {e}")
            QMessageBox.critical(self, self._("export to excel"),
                                 self._("Failed to export: ") + str(e))

    # ─────────────────────────────────────────────────────────────────────
    # I18N
    # ─────────────────────────────────────────────────────────────────────

    def retranslate_ui(self):
        try:
            self._ = TranslationManager.get_instance().translate
            self.btn_add.setText(self._("add"))
            self.btn_export.setText(self._("export to excel"))
            self.btn_refresh.setText(self._("refresh"))
            self.chk_admin_cols.setText(self._("show_admin_columns"))
            self.search_bar.setPlaceholderText(self._("search") + "...")
            self._lbl_rows_per_page.setText(self._("rows_per_page"))
            self._lbl_empty_text.setText(self._("no_data_available"))
            self._update_pagination_label()
            # هيدر الجدول
            if self.columns and self.table.columnCount() == len(self.columns):
                for i, col in enumerate(self.columns):
                    h = self.table.horizontalHeaderItem(i)
                    if h:
                        h.setText(self._(col.get("label", "")) if col.get("label") else "")
        except Exception as e:
            logger.warning(f"BaseTab retranslate_ui failed: {e}")

    # ─────────────────────────────────────────────────────────────────────
    # BACKWARD COMPATIBILITY
    # ─────────────────────────────────────────────────────────────────────
    # حفظ أسماء قديمة تستخدمها التابات الفرعية الموجودة

    def setup_ui(self):          self._setup_ui()
    def setup_table(self):       self._setup_table()
    def setup_shortcuts(self):   self._setup_shortcuts()
    def setup_signals(self):     self._setup_signals()
    def stretch_all_columns(self): self._stretch_columns()
    def show_context_menu(self, pos): self._show_context_menu(pos)

    @property
    def layout(self):
        """
        backward-compat: self.layout.addWidget(...) يرجع self._layout.
        التابات التي تحتاج Qt layout() الأصلية تستخدم super().layout().
        """
        return self._layout

    @property
    def filter_box(self):
        """
        backward-compat: بعض التابات (transactions_tab) تُخفيه.
        نُرجع QComboBox مخفي مشترك.
        """
        if not hasattr(self, "_filter_box_compat"):
            self._filter_box_compat = QComboBox()
            self._filter_box_compat.setVisible(False)
        return self._filter_box_compat

    def apply_settings(self):
        """backward-compat — الثيم يُطبَّق من ThemeManager مركزياً."""
        pass